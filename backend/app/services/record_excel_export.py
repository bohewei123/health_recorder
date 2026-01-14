from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _date_range(start_date: str, end_date: str) -> list[str]:
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    if end < start:
        start, end = end, start
    dates = []
    cur = start
    while cur <= end:
        dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return dates


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def build_health_records_workbook(
    start_date: str,
    end_date: str,
    records: list[dict],
    summaries_by_date: dict[str, dict],
) -> bytes:
    dates = _date_range(start_date, end_date)
    time_slots = ["起床", "上午", "下午", "晚上"]

    records_by_date_time: dict[str, dict[str, dict]] = {}
    for r in records:
        d = r.get("date")
        t = r.get("time_of_day")
        if not d or not t:
            continue
        records_by_date_time.setdefault(d, {})
        if t not in records_by_date_time[d]:
            records_by_date_time[d][t] = r

    wb = Workbook()
    ws = wb.active
    ws.title = "身体感觉和活动日志"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True)

    title_fill = PatternFill("solid", fgColor="F4B183")
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    label_fill = PatternFill("solid", fgColor="FCE4D6")

    total_cols = 1 + len(dates) * len(time_slots)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value="身体感觉和活动日志").font = title_font
    ws.cell(row=1, column=1).alignment = center
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = title_fill
        cell.border = border

    ws.cell(row=2, column=1, value="日期").font = header_font
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).border = border

    ws.cell(row=3, column=1, value="时间").font = header_font
    ws.cell(row=3, column=1).alignment = center
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=3, column=1).border = border

    col = 2
    for d in dates:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 3)
        date_cell = ws.cell(row=2, column=col, value=d)
        date_cell.font = header_font
        date_cell.alignment = center
        for c in range(col, col + 4):
            cell = ws.cell(row=2, column=c)
            cell.fill = header_fill
            cell.border = border

        for i, slot in enumerate(time_slots):
            cell = ws.cell(row=3, column=col + i, value=slot)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        col += 4

    layout_rows = [
        ("疼痛感觉（0～10）", "slot", "pain_level"),
        ("头晕感觉（0～10）", "slot", "dizziness_level"),
        ("情绪状态（0～10）", "slot", "mood_level"),
        ("描述身体感觉", "slot", "body_feeling_note"),
        ("前夜睡眠情况", "date", "sleep_note"),
        ("当日身体活动情况", "date", "daily_activity_note"),
        ("加重疼痛的活动", "date", "pain_increasing_activities"),
        ("减轻疼痛的活动", "date", "pain_decreasing_activities"),
        ("加重头晕的活动", "date", "dizziness_increasing_activities"),
        ("减轻头晕的活动", "date", "dizziness_decreasing_activities"),
        ("是否使用药物", "date", "medication_used"),
        ("用药说明", "date", "medication_note"),
    ]

    start_row_idx = 4
    for r_i, (label, row_type, field) in enumerate(layout_rows):
        row_idx = start_row_idx + r_i
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = header_font
        label_cell.fill = label_fill
        label_cell.alignment = center
        label_cell.border = border

        col = 2
        for d in dates:
            if row_type == "date":
                ws.merge_cells(start_row=row_idx, start_column=col, end_row=row_idx, end_column=col + 3)
                summary = summaries_by_date.get(d) or {}
                value = summary.get(field, "")
                if field == "medication_used":
                    value = "是" if bool(value) else "否"
                cell = ws.cell(row=row_idx, column=col, value=_stringify(value))
                cell.alignment = left_wrap if field.endswith("_note") or "activities" in field else center
                for c in range(col, col + 4):
                    sc = ws.cell(row=row_idx, column=c)
                    sc.border = border
                col += 4
                continue

            for i, slot in enumerate(time_slots):
                record = records_by_date_time.get(d, {}).get(slot)
                value = ""
                if record:
                    value = record.get(field, "")
                cell = ws.cell(row=row_idx, column=col + i, value=_stringify(value))
                cell.alignment = left_wrap if field.endswith("_note") else center
                cell.border = border
            col += 4

        for c in range(2, total_cols + 1):
            ws.cell(row=row_idx, column=c).border = border

    ws.column_dimensions["A"].width = 18
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    for r in range(start_row_idx, start_row_idx + len(layout_rows)):
        ws.row_dimensions[r].height = 48

    for r in range(1, start_row_idx + len(layout_rows)):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.alignment is None:
                cell.alignment = center

    ws2 = wb.create_sheet("明细")
    headers = [
        "id",
        "date",
        "time_of_day",
        "pain_level",
        "dizziness_level",
        "mood_level",
        "body_feeling_note",
        "stomach_level",
        "throat_level",
        "dry_eye_level",
        "fatigue_level",
        "sleep_note",
        "daily_activity_note",
        "pain_increasing_activities",
        "pain_decreasing_activities",
        "dizziness_increasing_activities",
        "dizziness_decreasing_activities",
        "medication_used",
        "medication_note",
        "notes",
        "triggers",
        "interventions",
        "created_at",
    ]

    for c, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r_i, record in enumerate(records, start=2):
        for c_i, h in enumerate(headers, start=1):
            v = record.get(h, "")
            if h == "medication_used":
                v = "是" if bool(v) else "否"
            cell = ws2.cell(row=r_i, column=c_i, value=_stringify(v))
            cell.alignment = left_wrap
            cell.border = border

    for c in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 18 if c <= 7 else 22
    ws2.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

